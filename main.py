import time
import openpyxl
import re
import os
import selenium
import datetime
import sys
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl.utils import column_index_from_string
import tkinter.filedialog as filedialog
import requests
import json
import requests
from bs4 import BeautifulSoup
from yahoofinancials import YahooFinancials
from forex_python.converter import CurrencyRates
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import threading
import tkinter.filedialog as filedialog
from datetime import datetime
import urllib3
from urllib3.exceptions import InsecureRequestWarning
import requests


def get_exchange_rate(base_currency, target_currency, exchange_rates):
    try:
        exchange_rate = exchange_rates[f"{base_currency}_{target_currency}"]
        return exchange_rate
    except KeyError:
        raise Exception(f"Failed to fetch exchange rate for {base_currency} to {target_currency}")


def scrape_product(driver, product_id, url_prefix, currency, exchange_rates):

    url = f"{url_prefix}{product_id}"
    driver.get(url)
    waittime=int(5)
    try:
        WebDriverWait(driver, waittime).until(EC.presence_of_element_located((By.CLASS_NAME, "pip-header-section")))
    except TimeoutException:
        return ["Product does not exist"] * 7  # Return a list with 7 "Product does not exist" values

    try:
        first_link = WebDriverWait(driver, waittime).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".pip-product-compact a"))).get_attribute("href")
    except TimeoutException:
        return ["Product does not exist"] * 7  # Return a list with 7 "Product does not exist" values

    driver.execute_script(f"window.open('{first_link}', '_blank')")
    driver.switch_to.window(driver.window_handles[-1])

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "pip-header-section")))
    product_name = driver.find_element_by_css_selector(".pip-header-section__title--big").text
    product_price = driver.find_element_by_css_selector(".pip-temp-price__integer").text
    product_price_str = re.sub(r'[^\d]', '', product_price)

    product_code = driver.find_element_by_css_selector(".pip-product-identifier__value").text
    product_description = driver.find_element_by_css_selector(".pip-header-section__description-text").text
    try:
        product_measurement_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".pip-header-section__description-measurement")))
        product_measurement = product_measurement_element.text
    except TimeoutException:
        product_measurement = "Not available"

    if currency == "PLN":
        product_price_pln = int(product_price_str)
        product_price_czk = int(product_price_str) * get_exchange_rate("PLN", "CZK", exchange_rates)
    elif currency == "CZK":
        product_price_czk = int(product_price_str)
        product_price_pln = int(product_price_str) * get_exchange_rate("CZK", "PLN", exchange_rates)

    driver.close()
    driver.switch_to.window(driver.window_handles[0])

    return [product_name, product_price_pln, product_price_czk, product_code, product_description, product_measurement, first_link]


def auto_adjust_columns(worksheet):
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2


def write_to_excel(workbook, data, sheet_name, currency):
    if not data:
        return

    if not workbook.sheetnames:
        sheet = workbook.active
        sheet.title = sheet_name
    else:
        sheet = workbook.create_sheet(sheet_name)

    headers = ["Product Name", "Product Price (PLN)", "Product Price (CZK)",
               "Product Code", "Product Description", "Product Measurement", "Link of the item"]
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_data in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num, value=cell_data)

    auto_adjust_columns(sheet)

def create_summary_sheet(workbook, data_cz, data_pl):
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    headers = ["Product Name", "Product Code", "Product Measurement", "Czech Price (CZK)", "Poland Price (CZK)"]
    for col_num, header in enumerate(headers, 1):
        summary_sheet.cell(row=1, column=col_num, value=header)

    for index, (cz_row, pl_row) in enumerate(zip(data_cz, data_pl), 2):
        product_name_cz, price_pln_cz, price_czk_cz, product_code_cz, product_description_cz, product_measurement_cz, link_cz = cz_row
        product_name_pl, price_pln_pl, price_czk_pl, product_code_pl, product_description_pl, product_measurement_pl, link_pl = pl_row

        if product_name_cz == product_name_pl:
            product_name = product_name_cz
        else:
            product_name = f"{product_name_cz} / {product_name_pl}"

        summary_data = [product_name, product_code_cz, product_measurement_cz, price_czk_cz, price_czk_pl]
        for col_num, cell_data in enumerate(summary_data, 1):
            summary_sheet.cell(row=index, column=col_num, value=cell_data)

    last_row = summary_sheet.max_row

    summary_sheet.cell(row=last_row + 1, column=1, value="Total:")
    summary_sheet.cell(row=last_row + 1, column=4, value=f"=SUM(D2:D{last_row})")
    summary_sheet.cell(row=last_row + 1, column=5, value=f"=SUM(E2:E{last_row})")

    # Add formatting for the total row
    for col_num in range(1, 6):
        summary_sheet.cell(row=last_row + 1, column=col_num).font = openpyxl.styles.Font(bold=True)

    auto_adjust_columns(summary_sheet)

def main(product_id_list, output_file_path, exchange_rates):
    current_dir = os.path.dirname(sys.executable)
    chrome_driver_path = os.path.join(current_dir, "chromedriver")
    driver = webdriver.Chrome(executable_path=chrome_driver_path)
    
    # Build the path to the chromedriver in the same directory as the main executable
    data_cz = []
    url_prefix_cz = "https://www.ikea.com/cz/cs/search/?q="
    target_currency = "CZK"
    for product_id in product_id_list:
        data_cz.append(scrape_product(driver, product_id, url_prefix_cz, target_currency, exchange_rates))

    data_pl = []
    url_prefix_pl = "https://www.ikea.com/pl/pl/search/?q="
    for product_id in product_id_list:
        data_pl.append(scrape_product(driver, product_id, url_prefix_pl, "PLN", exchange_rates))
    
    driver.quit()

    workbook = openpyxl.Workbook()
    write_to_excel(workbook, data_cz, "Czech Data", "CZK")
    write_to_excel(workbook, data_pl, "Poland Data", "PLN")
    create_summary_sheet(workbook, data_cz, data_pl)

    workbook.save(output_file_path)
    return data_cz, data_pl

def start_scraping():
    product_ids = product_ids_entry.get()
    product_id_list = [x.strip() for x in product_ids.split(",")]

    if not product_id_list:
        messagebox.showerror("Error", "Please enter the product IDs.")
        return

    if not output_folder_var.get():
        messagebox.showerror("Error", "Please choose the output folder.")
        return

    start_button.config(state=tk.DISABLED)
    progress_label.config(text="Scraping...")

    def run_scraping():
        try:
            exchange_rates = {
                "CZK_PLN": float(exchange_rate_czk_pln_var.get()),
                "PLN_CZK": float(exchange_rate_pln_czk_var.get())
            }

            output_file_path = get_output_file_path()  # Get the output file path
            data_cz, data_pl = main(product_id_list, output_file_path, exchange_rates)  # Pass output_file_path and exchange_rates to main()
            summary = f"Scraped {len(data_cz)} Czech and {len(data_pl)} Poland products."
            progress_label.config(text=summary)
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
            progress_label.config(text="")
        finally:
            start_button.config(state=tk.NORMAL)

    threading.Thread(target=run_scraping).start()


def browse_output_folder():
    output_folder = filedialog.askdirectory()
    output_folder_var.set(output_folder)

def get_output_file_path():
    current_datetime = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_folder = output_folder_var.get()
    output_file_name = f"ikea_products_{current_datetime}.xlsx"
    return os.path.join(output_folder, output_file_name)

app = tk.Tk()
app.title("IKEA Product Scraper")

frame = ttk.Frame(app, padding="10")
frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

exchange_rate_czk_pln_label = ttk.Label(frame, text="Exchange rate CZK-PLN:")
exchange_rate_czk_pln_label.grid(row=4, column=0, padx=(0, 10), pady=(0, 10), sticky=tk.W)

exchange_rate_czk_pln_var = tk.StringVar()
exchange_rate_czk_pln_entry = ttk.Entry(frame, width=10, textvariable=exchange_rate_czk_pln_var)
exchange_rate_czk_pln_entry.insert(0, "0.1916")  # Default value
exchange_rate_czk_pln_entry.grid(row=4, column=1, padx=(0, 10), pady=(0, 10), sticky=tk.W)

exchange_rate_pln_czk_label = ttk.Label(frame, text="Exchange rate PLN-CZK:")
exchange_rate_pln_czk_label.grid(row=5, column=0, padx=(0, 10), pady=(0, 10), sticky=tk.W)

exchange_rate_pln_czk_var = tk.StringVar()
exchange_rate_pln_czk_entry = ttk.Entry(frame, width=10, textvariable=exchange_rate_pln_czk_var)
exchange_rate_pln_czk_entry.insert(0, "5.21")  # Default value
exchange_rate_pln_czk_entry.grid(row=5, column=1, padx=(0, 10), pady=(0, 10), sticky=tk.W)


product_ids_label = ttk.Label(frame, text="Product IDs (comma separated):")
product_ids_label.grid(row=0, column=0, padx=(0, 10), pady=(0, 10), sticky=tk.W)

product_ids_entry = ttk.Entry(frame, width=60)
product_ids_entry.insert(0, "703.780.70,194.311.70,694.780.23")
product_ids_entry.grid(row=0, column=1, padx=(0, 10), pady=(0, 10), sticky=tk.W)

output_folder_label = ttk.Label(frame, text="Output folder:")
output_folder_label.grid(row=2, column=0, padx=(0, 10), pady=(0, 10), sticky=tk.W)

output_folder_var = tk.StringVar()
output_folder_entry = ttk.Entry(frame, width=60, textvariable=output_folder_var)
output_folder_entry.insert(0, "/Users/MarekHalska/Downloads")
output_folder_entry.grid(row=2, column=1, padx=(0, 10), pady=(0, 10), sticky=tk.W)

browse_button = ttk.Button(frame, text="Browse", command=browse_output_folder, padding=(5, 0))
browse_button.grid(row=2, column=2, padx=(0, 10), pady=(0, 10), sticky=tk.W)

start_button = ttk.Button(frame, text="Start Scraping", command=start_scraping)
start_button.grid(row=6, column=0, columnspan=3, pady=(0, 10))

progress_label = ttk.Label(frame, text="", wraplength=300)
progress_label.grid(row=7, column=0, columnspan=3)

app.mainloop()

